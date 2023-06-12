from openpyxl import load_workbook


class ResultsImporter(object):
    def __init__(self, ranker):
        self._ranker = ranker

        # Temporary variables set during import
        self._race_info = {
            "Race Name": None,
            "Race Date": None,
            "Race Length": None,
            "Winning Time": None,
            "City": None,
            "State": None,
        }

        # Require Columns Locations
        self._column_mapping = {
            "Team Name": None,
            "Division": None,
            "Overall Place": None,
            "Division Place": None,
            "Racer 1": None,
            "Racer 2": None,
            "Racer 3": None,
            "Racer 4": None,
        }

    def import_file(self, path):
        print("Importing ", path)

        if path.suffix == ".xlsx":
            wb = load_workbook(path)

            required = ["Instructions", "Race Info", "Results"]
            for r in required:
                if r not in wb.sheetnames:
                    raise ValueError(f"Missing {r} worksheet: {path}")

            self._import_race_info(wb.get_sheet_by_name("Race Info"))
            self._import_results(wb.get_sheet_by_name("Results"))
        else:
            raise ValueError("Unknown file type: ", path)

    def _import_race_info(self, race_info_sheet):
        # Assume the Race Info field names are the first column and the
        # user-provided data is the second column.
        for cell in race_info_sheet["A"]:
            if cell.value in self._race_info:
                self._race_info[cell.value] = cell.offset(column=1).value

        # Allow winning time to be missing
        if self._race_info["Winning Time"] is None:
            self._race_info["Winning Time"] = "?"

        # Verify we found the required Race Info fields.
        if None in self._race_info.values():
            raise ValueError("Missing some race info:", self._race_info)

    def _import_results(self, results_sheet):
        # Get column indexes of required columns from the first row, this allows
        # them to be in any order.
        for cnum in range(1, results_sheet.max_column + 1):
            field = results_sheet.cell(row=1, column=cnum).value
            if field in self._column_mapping:
                self._column_mapping[field] = cnum

        # Verify we found the required columns.
        if None in self._column_mapping.values():
            raise ValueError("Missing some results columns:", self._column_mapping)

        def cell_to_var(rnum, field):
            """Just a quick lambda to reduce some code in the next loop."""
            cell = results_sheet.cell(row=rnum, column=self._column_mapping[field])
            return cell.value

        # Actual data starts on 2nd row
        for rnum in range(2, results_sheet.max_row + 1):
            team_name = cell_to_var(rnum, "Team Name")
            division = cell_to_var(rnum, "Division")
            overall_place = cell_to_var(rnum, "Overall Place")
            division_place = cell_to_var(rnum, "Division Place")

            members = []
            for x in range(1, 5):  # Racers 1-4
                racer = cell_to_var(rnum, f"Racer {x}")
                if racer is not None:
                    members.append(racer)

            if division is None:
                return

            self._ranker.add_entry(
                self._race_info,
                division,
                team_name,
                overall_place,
                division_place,
                members,
            )
